# -*- coding: utf-8 -*-

# Maak een importbestand voor de centrumsteden: alle itemcodes van de connector, maar volgens
#   de indeling van de externe PinC-themaboom
#   Gecreëerd op 2025-02-06 door Geert Bonte
#   Laatst gewijzigd op 2025-08-25

# Voorbereiding:
#   Exporteer de volledige PinC-themaboom (CategoryTree.xlsx) vanuit Swing Studio
#
# De output van dit script is een themaboom die geïmporteerd kan worden bij de centrumsteden: CategoryTree_import.xlsx.

import openpyxl
import warnings

from datetime import datetime
from os import getenv
from os.path import exists, getctime

# initialiseer enkele variabelen
pinc_extern = ["Thema's", 'PRODUCTIE', 'EXTERN']  # de externe themaboom van PinC
cs_connector = ["Thema's", 'PRODUCTIE', 'INTERN', 'Swing Connectoren', 'Uitgaande connectoren', 'Centumsteden', 'Swing Connector Centrumsteden']  # de uitgaande connector naar de centrumsteden
dna_import = ['import_PinC']  # importmap voor de centrumsteden
file_out = 'CategoryTree_import_PinC.xlsx'
themes = []  # zal een lijst van hoofdthema's bevatten

# return a list of cell values for a worksheet row
def cellvalues(row):
    values = []
    for cell in row:
        values.append(cell.value)
    return values

# fix workbook, worksheet and save to Excel file
def fix_and_save_excel(wb, ws, file):
    # pas de kolombreedtes aan
    col_no = 0
    for col in ws.columns:
        col_no = col_no + 1
        set_len = 0
        column = col[0].column_letter  # get the column name
        row = 0
        for cell in col:
            char_len = len(str(cell.value))
            if char_len > set_len:
                set_len = char_len
            row += 1
        ws.column_dimensions[column].width = set_len + 3  # set column width to maximal cell length
        ws.auto_filter.ref = ws.dimensions  # autofilter toevoegen
    ws.freeze_panes = 'B2'
    # wijzig auteur van 'openpyxl' naar de huidige Windows-gebruiker
    wb.properties.creator = getenv('USERNAME')
    # wegschrijven
    print(f'Wegschrijven naar {file}')
    wb.save(file)


# hoofdprogramma
def main():

    # lees themaboom
    category_tree = 'CategoryTree.xlsx'
    if exists(category_tree):
        dt = str(datetime.fromtimestamp(getctime(category_tree)))[:10]  # creatiedatum van het Excel-bestand
        print(f'Bezig met lezen van themaboom ({category_tree} – geëxporteerd op {dt})')
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                wb = openpyxl.load_workbook(category_tree)
        except:
            print(f'\nKan {category_tree} niet lezen. Het bestand is wellicht geopend in Excel of een andere toepassing. Sluit het bestand en run opnieuw.')
            return
        ws_cat = wb.active
    else:
        print(f'\n{category_tree} niet gevonden. Exporteer de themaboom en de indicatortabel met Studio en plaats de geëxporteerde bestanden hier.')
        return

    # verwerk themaboom
    # kolom A: ItemCode: indicatorcode
    # kolom B: CategoryRoot: "Thema's"
    # kolom C: CategoryName: "PRODUCTIE" | "ADMIN ONLY"
    # kolom D: CategoryName: "EXTERN" | "INTERN" | ...
    # kolom E e.v.: CategoryName: (thema's en subthema's)
    max_col = len(ws_cat[1])  # maximaal aantal kolommen
    pinc_max_col = len(pinc_extern)  # aantal kolommen in pinc_extern (externe themaboom PinC)
    cs_max_col = len(cs_connector)  # aantal kolommen in cs_connector (uitgaande connector)
    dna_max_col = len(dna_import)  # aantal kolommen in dna_import (import voor centrumsteden)

    # stap 1: maak een lijst met alle itemcodes onder de connector
    unique_vars = {}  # key = subtree, value = dictionary (key = onderwerpcode, value = 1)
    n = 0
    for row in ws_cat:
        if n > 0:
            values = cellvalues(row)
            if values[1:cs_max_col+1] == cs_connector:
                unique_vars[values[0]] = 1
        n += 1
    unique_vars = unique_vars.keys()  # lijst van unieke itemcodes in de connector
    
    # stap 2: creëer de import-themaboom: overloop de externe themaboom van PinC en verwijder alle itemcodes die niet in de connector zitten
    wb_import = openpyxl.Workbook()
    ws_import = wb_import.active
    ws_import.title = 'D&A import'
    # kopieer de rij met kolomkoppen
    for c in range(0, max_col):
        column_letter = chr(ord('A') + c)
        ws_import[column_letter + '1'] = ws_cat[column_letter + '1'].value
    # kopieer derijen met data
    n = 0  # rijnummer in ws_cat
    i = 1  # rijnummer in ws_import
    unique_folders = []
    m = 0
    for row in ws_cat:
        if n > 0:
            values = cellvalues(row)
            if values[1:pinc_max_col+1] == pinc_extern:
                itemcode = values[0]
                if itemcode in unique_vars:
                    theme = values[4]
                    if not theme in themes:
                        themes.append(theme)
                    folders = values[1:]
                    if not folders in unique_folders:
                        unique_folders.append(folders)
                        m += 1
                    i += 1
                    values[0] = 'dna_' + values[0]
                    ws_import['A' + str(i)] = values[0]
                    for c in range(0, dna_max_col):
                        column_letter = chr(ord('A') + c + 1)
                        ws_import[column_letter + str(i)] = dna_import[c]
                    for c in range(pinc_max_col + 1, len(values)):
                        if values[c]:
                            column_letter = chr(ord('A') + c - pinc_max_col + dna_max_col)
                            ws_import[column_letter + str(i)] = values[c]
        n += 1
    ws_import['B1'].value = 'CategoryName'  # vervang CategoryRoot door CategoryName
    print('{} items (onderwerpen, rapporten, presentaties en URL-links)'.format(f'{n:,d}'.replace(',','.')))
    print('{} mappen en submappen'.format(f'{m:,d}'.replace(',','.')))

    # fix de lay-out en schrijf weg naar Excel-bestand
    fix_and_save_excel(wb_import, ws_import, file_out)
    print('  {} items'.format(f'{n:,d}'.replace(',','.')))
    
    # maak aparte bestanden aan per hoofdthema
    for theme in themes:
        theme_out = file_out.replace('.', f'_{theme}.').replace(' ', '_').replace(',', '')
        wb_theme = openpyxl.Workbook()
        ws_theme = wb_theme.active
        ws_theme.title = f'D&A import {theme}'[:31]  # werkbladnaam mag maximaal 31 tekens lang zijn
        # kopieer de rijen met data
        n = 0  # rijnummer in ws_theme
        for row in ws_import:
            values = cellvalues(row)
            if n == 0 or values[2] == theme:
                n += 1
                for c in range(0, len(values)):
                    column_letter = chr(ord('A') + c)
                    ws_theme[column_letter + str(n)] = values[c]                
        # fix de lay-out en schrijf weg naar Excel-bestand
        fix_and_save_excel(wb_theme, ws_theme, theme_out)
        print('  {} items'.format(f'{n:,d}'.replace(',','.')))

    return  # einde programma

# hoofdprogramma aanroepen als een functie
if __name__ == '__main__':
    main()
