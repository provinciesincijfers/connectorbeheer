# -*- coding: utf-8 -*-

# Maak een overzicht van de themaboom van Swing met alle indicatoren
#   Gecreëerd op 2024-05-27 door Geert Bonte
#   Laatst gewijzigd op 2025-02-25

# Voorbereiding:
#   1. Check het configuratiebestand (swing_sites.xlsx) en zorg dat de rij die betrekking heeft op jouw Swing-site
#      bovenaan staat (onder de rij met de kolomhoofden).
#   2. Exporteer eerst de themaboom (CategoryTree.xlsx) en de indicatortabel (Indicators.xlsx) vanuit Swing Studio.
#      Er kunnen immers wijzigingen gebeurd zijn sinds je dit script laatst hebt gerund.
#      Het is belangrijk dat de themaboom en de indicatortabel op hetzelfde moment geëxporteerd worden.
#   3. Optioneel: exporteer ook de rapportentabel (Reports.xlsx) en de URL-links (Url_links.xlsx) om de links en
#      de rapporten op te nemen als aanklikbare links (anders blijven er weinig zeggende itemcodes staan).
#
# De output van dit script is één of meer HTML-bestanden, die je met een webbrowser kunt openen en vervolgens kunt
#   opslaan als PDF. Het script produceert deze bestanden, afhankelijk van welk deel van de themaboom je uit Studio
#   geëxporteerd hebt (*):
#     themaboom_extern.html
#     themaboom_intern.html
#     themaboom_admin.html
#     themaboom_unused.html
#
# (*) Wanneer je de volledige themaboom hebt geëxporteerd, krijg je een outputbestand voor elk van de mogelijke
#     deelthemabomen (extern, intern, admin). Je kunt ook enkel de externe, enkel de interne of enkel de admin-
#     themaboom exporteren, dan krijg je enkel het relevante outputbestand.
#
#     Als de geëxporteerde indicatortabel onderwerpen bevat die op geen enkele plaats in één of meer van de themabomen
#     staan, dan wordt er een extra outputbestand aangemaakt met een overzicht van die onderwerpen.
#
# Alle thema’s en onderwerpen zijn aanklikbare links naar de databank.

import io
import openpyxl
import re
import warnings

from datetime import date, datetime
from os.path import exists, getctime
from urllib.parse import quote

# initialiseer enkele variabelen
sites_file = 'swing_sites.xlsx'
config = {  # default configuratie voor het geval het configuratiebestand (swing_sites.xlsx) ontbreekt
    'site_name': 'Provincies in cijfers',
    'database_url': 'https://provincies.incijfers.be/databank',
    'categories': {
        'extern': ['PRODUCTIE', 'EXTERN'],
        'intern': ['PRODUCTIE', 'INTERN'],
        'admin': ['ADMIN ONLY']
    },
    'show_description': '',
    'subtrees_count': 3
}

# Het configuratiebestand is een Excel-bestand (swing_sites.xlsx) met volgende kolommen:
#   site_owner, site_name, database_url, cat_extern, cat_intern, cat_admin
# We kijken enkel naar de eerste datarij. In de kolommen cat_extern, cat_intern en cat_admin vullen we de
#   namen in van de deelthemabomen (met een komma ertussen indien 2 niveaus).
# Voorbeelden:
#   site_owner site_name              database_url                              cat_extern         cat_intern         cat_admin
#   PinC       Provincies in cijfers  https://provincies.incijfers.be/databank  PRODUCTIE, EXTERN  PRODUCTIE, INTERN  ADMIN ONLY
#   Kortrijk   Kortrijk in cijfers    https://kortrijk.incijfers.be/jive        Extern             Intern    
    

# bepaal voor elke rij van de geëxporteerde themaboom tot welke themaboom de rij behoort (extern, intern of admin)
def get_subtree(row):
    subtree = ''
    categories = config['categories']
    for key in categories:
        if not subtree:
            category = categories[key]
            n = len(category)
            subtree = key
            for c in range(0, n):
                if row[c+2].value != category[c]: subtree = ''
    #if not subtree:
    #    print(f"\nOnbekende themaboom: '{row[c+2].value}'")
    return subtree

# HTML encode (thema en onderwerpnaam in HTML-tekst)
def htmlencode(text):
    text = text.replace('&', '&amp;')
    text = text.replace('<', '&lt;')
    text = text.replace('>', '&gt;')
    text = text.replace("'", '&apos;')
    text = text.replace('"', '&quot;')
    return text

# lees Excel-configuratiebestand (swing_sites.xlsx)
def read_sites():
    global config, sites_file
    
    # lees Excel-bestand als het bestaat
    if not exists(sites_file):
        return False
    else:
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                wb = openpyxl.load_workbook(sites_file)
        except:
            print(f'\nKan {sites_file} niet lezen. Het bestand is wellicht geopend in Excel of een andere toepassing. Sluit het bestand en run opnieuw.')
            return False
        ws_cat = wb.active
        # verwerk sites: lees de rij met de kolomhoofden en lees enkel de eerste datarij
        # kolommen: site_owner, site_name, database_url, cat_extern, cat_intern, cat_admin, show_description
        fields = []
        config['categories'] = {}
        config['show_description'] = ''
        config['subtrees_count'] = 0
        for col_no in range(1, ws_cat.max_column + 1):  # lees de eerste rij (kolomhoofden)
            field = ws_cat.cell(row=1, column=col_no).value.lower()
            fields.append(field)
        for col_no in range(1, ws_cat.max_column + 1):  # lees de tweede rij (data)
            field = fields[col_no-1]
            val = ws_cat.cell(row=2, column=col_no).value or ''
            if field in ['site_name', 'database_url']:
                if not val:
                    print(f'\nOntbrekende waarden in {sites_file}. Pas het bestand aan en run opnieuw.')
                    return False
                config[field] = val
            elif field.startswith('cat_'):
                cat = field[4:]
                val = re.sub(', +', ',', val)
                if val: config['categories'][cat] = val.split(',')
            elif field == 'show_description' and val:
                config['show_description'] = 'x'
        if config['categories'] == {}:
            print(f'\nOntbrekende waarden in {sites_file}. Pas het bestand aan en run opnieuw.')
            return False
        config['subtrees_count'] = len(config['categories'])
        return True

# URL encode (thema in URL-querystringparameter)
def urlencode(text):
    text = quote(text)
    text = text.replace('/', '%2F')
    return text

# markeer speciale indicatornamen
def itemname(item_name):
    item_name = item_name.replace('ARCHIEF', '<span class="warn">ARCHIEF</span>')
    item_name = item_name.replace('GEPLAND', '<span class="warn">GEPLAND</span>')
    item_name = item_name.replace('UITDOVEND', '<span class="warn">UITDOVEND</span>')
    item_name = item_name.replace('UITGEDOOFD', '<span class="warn">UITGEDOOFD</span>')
    item_name = item_name.replace('WEG', '<span class="warn">WEG</span>')
    item_name = re.sub(r'(</span>) ?(\([^\)]+?\))', r' \2\1 ', item_name).replace('  ', ' ')
    return item_name

# hoofdprogramma
def main():
    global config, sites_file
    if not read_sites(): return  # lees het Excel-configuratiebestand, ga enkel verder als bestand oké is
    
    variables = {}  # key = onderwerpcode, value = tuple: (onderwerpnaam, thema)
    reports = {}  # key = rapportcode, value = tuple: (rapportnaam, bestandsnaam, type)
    urllinks = {}  # key = urllinkcode, value = tuple: (urllinknaam, omschrijving, url)
    site_name = config['site_name']
    base_url = config['database_url']
    site_link = f'<a href="{base_url}" target="_blank">{site_name}</a>'

    print('Website:', site_name)
    print('URL van de databank:', base_url)

    today = date.today()
    html = {}  # key = subtree, value = html
    count = {}  # key = subtree, value = list(thema's, onderwerpen)
    unique_vars = {}  # key = subtree, value = dictionary (key = onderwerpcode, value = 1)
    export_dates = { 'cat': '', 'ind': '', 'rpt': '', 'urls': '' }

    # (optioneel) lees rapportentabel
    reportsfile = 'Reports.xlsx'
    if exists(reportsfile):
        dt = str(datetime.fromtimestamp(getctime(reportsfile)))[:10]  # creatiedatum van het Excel-bestand
        print(f'Bezig met lezen van rapportentabel ({reportsfile} – geëxporteerd op {dt})')
        export_dates['rpt'] = dt
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                wb = openpyxl.load_workbook(reportsfile)
        except:
            print(f'\nKan {reportsfile} niet lezen. Het bestand is wellicht geopend in Excel of een andere toepassing. Sluit het bestand en run opnieuw.')
            return
        ws_rpt = wb.active
        # verwerk rapportentabel
        # kolom A: rapportcode    row[0]
        # kolom B: rapportnaam    row[1]
        # kolom D: bestandsnaam   row[3]
        # kolom E: bestandstype   row[4]
        n = 0
        for row in ws_rpt:
            if n > 0:
                item_code = row[0].value
                report_name = row[1].value
                report_file = row[3].value
                report_type = row[4].value
                reports[item_code] = [report_name, report_file, report_type]
            n += 1

    # (optioneel) lees url-links
    urllinksfile = 'Url_links.xlsx'
    if exists(urllinksfile):
        dt = str(datetime.fromtimestamp(getctime(urllinksfile)))[:10]  # creatiedatum van het Excel-bestand
        print(f'Bezig met lezen van URL-linkstabel ({urllinksfile} – geëxporteerd op {dt})')
        export_dates['urls'] = dt
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                wb = openpyxl.load_workbook(urllinksfile)
        except:
            print(f'\nKan {urllinksfile} niet lezen. Het bestand is wellicht geopend in Excel of een andere toepassing. Sluit het bestand en run opnieuw.')
            return
        ws_urls = wb.active
        # verwerk urllinkstabel
        # kolom A: urllinkcode          row[0]
        # kolom B: urllinknaam          row[1]
        # kolom C: urllinkomschrijving  row[2]
        # kolom D: urllink-URL          row[3]
        n = 0
        for row in ws_urls:
            if n > 0:
                item_code = row[0].value
                urllink_name = row[1].value
                urllink_desc = row[2].value
                urllink_url  = row[3].value
                urllinks[item_code] = [urllink_name, urllink_desc, urllink_url]
            n += 1
    # lees themaboom
    category_tree = 'CategoryTree.xlsx'
    if exists(category_tree):
        dt = str(datetime.fromtimestamp(getctime(category_tree)))[:10]  # creatiedatum van het Excel-bestand
        print(f'Bezig met lezen van themaboom ({category_tree} – geëxporteerd op {dt})')
        export_dates['cat'] = dt
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                wb = openpyxl.load_workbook(category_tree)
        except:
            print(f'\nKan {category_tree} niet lezen. Het bestand is wellicht geopend in Excel of een andere toepassing. Sluit het bestand en run opnieuw.')
            return
        ws_cat = wb.active
    else:
        print(f'\n{category_tree} niet gevonden. Exporteer de themaboom en de indicatortabel met Studio en plaats de geëxporteerde bestanden hier.')
        return

    # lees indicatortabel
    indicators = 'Indicators.xlsx'
    if exists(indicators):
        dt = str(datetime.fromtimestamp(getctime(indicators)))[:10]  # creatiedatum van het Excel-bestand
        print(f'Bezig met lezen van indicatortabel ({indicators} – geëxporteerd op {dt})')
        export_dates['ind'] = dt
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                wb = openpyxl.load_workbook(indicators)
        except:
            print(f'\nKan {indicators} niet lezen. Het bestand is wellicht geopend in Excel of een andere toepassing. Sluit het bestand en run opnieuw.')
            return
        ws_ind = wb.active
    else:
        print(f'\n{indicators} niet gevonden. Exporteer de themaboom en de indicatortabel met Studio en plaats de geëxporteerde bestanden hier.')
        return
    
    # verwerk indicatortabel
    # kolom A: onderwerpcode  row[0]
    # kolom B: onderwerpnaam  row[1]
    # kolom C: omschrijving   row[2]
    # kolom N: datatype       row[13]
    # kolom P: geotype        row[15]
    # kolom AE: cube          row[30]
    n = 0
    count_cubes = 0  # totaal aantal kubusonderwerpen in de databank
    for row in ws_ind:
        if n > 0:
            item_code = row[0].value
            var_name = row[1].value
            var_desc = ''
            if config['show_description'] == 'x': var_desc = htmlencode(row[2].value)
            var_type = 'default'
            var_unused = True
            if row[13].value == 'Enum': var_type = 'label'
            if row[13].value == 'Mean': var_type = 'mean'
            if 'Percentage' in row[13].value: var_type = 'percent'
            if row[13].value == 'Html' or row[13].value == 'Text': var_type = 'text'
            if 'ratio' in row[4].value or 'promille' in row[4].value: var_type = 'ratio'
            if row[15].value == 'Stream': var_type = 'stream'
            if str(row[30].value) == '1':
                var_type = 'cube_' + var_type
                count_cubes += 1
            variables[item_code] = [var_name, var_desc, var_type, var_unused]
        n += 1
    count_total = n  # totaal aantal onderwerpen in de databank

    # verwerk themaboom
    # kolom A: ItemCode: indicatorcode
    # kolom B: CategoryRoot: "Thema's"
    # kolom C: CategoryName: "PRODUCTIE" | "ADMIN ONLY"
    # kolom D: CategoryName: "EXTERN" | "INTERN" | ...
    # kolom E e.v.: CategoryName: (thema's en subthema's)
    #   thema kan naast (accent)letters, cijfers en spaties ook volgende leestekens bevatten: '()/-&+,.:>[]₂
    #   om themanaam op te nemen als URL-parameter gebruiken we daarom urlencode() (eigen functie)
    max_col = len(ws_cat[1])  # maximaal aantal kolommen
    old_cat_path = ''
    themes = {}  # key = niveau, value = subthema
    old_themes = []  # index = niveau, value = subthema
    for c in range(0, max_col + 2):
        old_themes.append('')
    
    # overloop de themaboom rij per rij
    n = 0
    for row in ws_cat:
        if n > 0:
            subtree = get_subtree(row)  # bepaal de subthemaboom (bv. extern, intern of admin)
            if subtree:
                start_col = 2 + len(config['categories'][subtree])  # bepaal het kolomnummer waar het hoofdthema start
                if not subtree in html:
                    html[subtree] = ''  # html-tekst
                    count[subtree] = [0, 0]  # aantal (sub)thema's, aantal onderwerpen
                    unique_vars[subtree] = {}  # key = onderwerpcode, value = 1
                # bepaal thema en subthema
                cat_path = row[start_col].value or ''
                cat_param = 'cat_open=' + urlencode(cat_path) or ''  # querystringparameter: cat_open
                level = 1
                item_level = 2
                themes[1] = row[start_col].value
                # daal af naar het laagste subthema
                for c in range(start_col + 1, max_col):
                    level += 1
                    if row[c].value:
                        cat_path += ' > ' + row[c].value
                        cat_param += '/' + urlencode(row[c].value)
                        themes[level] = row[c].value
                        item_level = level + 1
                    else:
                        themes[level] = ''
                # voeg subthematitel(s) toe wanneer we naar een ander subthema overschakelen
                if cat_path != old_cat_path:
                    cat_url = f'{base_url}?{cat_param}'
                    for lvl in range(1, level + 1):
                        if themes[lvl] != old_themes[lvl]:
                            if themes[lvl]:
                                theme_name = htmlencode(themes[lvl])
                                # voeg subthematitel toe aan HTML-bestand
                                html[subtree] += f'\t\t<p class="theme level{lvl}"><a href="{cat_url}" target="category_tree">{theme_name}</a></p>\n'
                                count[subtree][0] += 1
                            old_themes[lvl] = themes[lvl]
                            old_themes[lvl + 1] = ''
                    print(subtree + ': ' + cat_path)
                old_cat_path = cat_path
                item_code = row[0].value
                if not cat_path: item_level = 1
                if item_code not in variables:  # item staat niet in de indicatortabel (is dus geen onderwerpcode)
                    # controleer of het een rapportcode is en zo ja, gebruik naam van rapport en link naar rapport
                    if item_code in reports:
                        report_name = reports[item_code][0]  # rapportnaam
                        report_file = reports[item_code][1]  # bestandsnaam
                        report_type = reports[item_code][2]  # bestandstype
                        report_name = itemname(htmlencode(report_name))
                        if report_type == 'html':
                            report_url = f'{base_url}/report?openinputs=true&id={item_code}'
                        else:  # pdf report
                            report_url = f'{base_url}/report?openinputs=true&id={item_code}'
                        print(f'  {report_name} [{item_code}]')
                        html[subtree] += f'\t\t<p class="report level{item_level}" title="{item_code}"><a href="{report_url}" target="_blank">{report_name}</a></p>\n'
                    # controleer of het een url-link-code is en zo ja, gebruik naam van de link en maak de link aanklikbaar
                    elif item_code in urllinks:
                        urllink_name = urllinks[item_code][0]  # urllinknaam
                        urllink_desc = urllinks[item_code][1]  # omschrijving
                        urllink_url  = urllinks[item_code][2]  # url
                        print(f'  {urllink_name} [{item_code}]')
                        html[subtree] += f'\t\t<p class="urllink level{item_level}" title="{urllink_desc}"><a href="{urllink_url}" target="_blank">{urllink_name}</a></p>\n'
                    else:
                        if reports and urllinks:
                            # itemcode kan enkel nog een presentatie zijn...
                            print(f'  (presentatie) [{item_code}]')
                            pres_url = f'{base_url}?presel_code={item_code}'
                            html[subtree] += f'\t\t<p class="pres level{item_level}" title="{item_code}"><a href="{pres_url}" target="_blank">{item_code}</a></p>\n'
                        else:
                            # we hebben geen idee wat dit is
                            print(f'  (presentatie) [{item_code}]')
                            html[subtree] += f'\t\t<p class="nonvar level{item_level}" title="{item_code}">{item_code}</p>\n'
                else:  # item staat in de indicatortabel (is dus een onderwerpcode)
                    var_name = variables[item_code][0]  # onderwerpnaam
                    var_desc = variables[item_code][1]  # omschrijving
                    var_type = variables[item_code][2]  # datatype
                    variables[item_code][3] = False  # to indicate this variable is used at least once in the category tree
                    var_param = 'var=' + item_code  # querystringparameter: var
                    var_url = f'{base_url}?{var_param}&{cat_param}'
                    print(f'  {var_name} [{item_code}]')
                    var_name = itemname(htmlencode(var_name))
                    # voeg onderwerp toe aan HTML-bestand
                    var_temp = ''
                    if 'temp' in item_code[0:4]: var_temp = 'temp'  # tijdelijk onderwerp
                    var_title = var_desc.replace('\n', '&#x0a;')
                    var_desc = var_desc.replace('\n', '<br/>')
                    html[subtree] += f'\t\t<p class="var level{item_level} {var_type}{var_temp}" title="{var_title}"><a href="{var_url}" target="category_tree">{var_name}</a><span class="desc">{var_desc}</span></p>\n'
                    count[subtree][1] += 1
                    unique_vars[subtree][item_code] = 1
        n += 1

    # rapporteer niet-gebruikte variabelen (niet in themaboom), alleen indien we alle deelthemabomen verwerkt hebben
    if len(html) == config['subtrees_count']:
        print('indicatoren die niet in de themaboom staan:')
        subtree = 'unused'
        html[subtree] = ''
        count[subtree] = [0, 0]  # aantal (sub)thema's, aantal onderwerpen
        unique_vars[subtree] = {}  # key = onderwerpcode, value = 1
        #variables = dict(sorted(variables.items()))  # sorteer op keys
        variables = dict(sorted(variables.items(), key=lambda x: x[1]))  # sorteer op values (omschrijving)
        for item_code in variables:
            var_name = variables[item_code][0]  # onderwerpnaam
            var_desc = variables[item_code][1]  # omschrijving
            var_type = variables[item_code][2]  # datatype
            var_unused = variables[item_code][3]  # used/unused flag
            var_param = 'var=' + item_code  # querystringparameter: var
            if var_unused == True:
                var_url = f'{base_url}?{var_param}'
                print(f'  {var_name} [{item_code}]')
                var_name = itemname(htmlencode(var_name))
                # voeg onderwerp toe aan HTML-bestand
                var_temp = ''
                if 'temp' in item_code[0:4]: var_temp = 'temp'  # tijdelijk onderwerp
                #var_title = var_desc.replace('\n', '&#x0a;')
                var_desc = var_desc.replace('\n', '<br/>')
                html[subtree] += f'\t\t<p class="var level2 {var_type}{var_temp}" title="{item_code}"><a href="{var_url}" target="category_tree">{var_name}</a><span class="desc">{var_desc}</span></p>\n'
                count[subtree][1] += 1
                unique_vars[subtree][item_code] = 1

    # schrijf HTML-bestanden
    print('\n')
    if not html.keys():
        print(f'Themaboom niet verwerkt. Check {sites_file} en run opnieuw.')
    else:
        for subtree in list(html.keys()):
            count_unique = len(unique_vars[subtree])
            file_out = f"themaboom_{subtree}_{export_dates['cat']}.html"
            if subtree == 'unused':
                print(f'Bezig met schrijven naar {file_out} ({count[subtree][1]:,d} onderwerpen)'.replace(',','.').replace(';',','))
            else:
                print(f'Bezig met schrijven naar {file_out} ({count[subtree][0]:,d} thema’s; {count[subtree][1]:,d} onderwerpen; {count_unique:,d} unieke onderwerpen)'.replace(',','.').replace(';',','))
            with io.open(file_out, 'w', encoding='utf8') as f:
                f.write(f"""<html>

    <head>
        <title>Themaboom {subtree} {export_dates['cat']} ({site_name})</title>
        <meta charset="UTF-8"/>
        <meta date="{today}"/>
        <link rel="stylesheet" href="swing_category_tree.css"/>
    </head>

    <body>

        <h1>Themaboom {subtree}</h1>
        
        <p class="date">{site_link} – {export_dates['cat']}</p>

""".replace('    ', '\t').replace('themaboom unused', 'Onderwerpen niet in themaboom'))
                if subtree == 'unused':
                    f.write('\t\t<p class="subtree">NIET IN THEMABOOM</p>\n')
                else:
                    subtree_title = ' '.join(config['categories'][subtree])
                    f.write(f'\t\t<p class="subtree">{subtree_title}</p>\n')
                f.write(html[subtree])
                f.write('\n')
                if subtree == 'unused':
                    f.write(f"\t\t<p class=\"count\">Op {export_dates['cat']} zijn {count[subtree][1]:,d} onderwerpen niet opgenomen in de themaboom.</p>\n".replace(',','.'))
                else:
                    f.write(f"\t\t<p class=\"count\">Op {export_dates['cat']} bevat de themaboom {subtree} {count[subtree][0]:,d} thema’s en subthema’s en {count[subtree][1]:,d} onderwerpen (waarvan {count_unique:,d} unieke onderwerpen).</p>\n".replace(',','.'))
                f.write(f'\t\t<p class="count">Er zijn in totaal {count_total:,d} onderwerpen in de databank (waarvan {count_cubes} kubusonderwerpen).</p>\n'.replace(',','.'))
                f.write('\n\t</body>\n\n</html>')

    # waarschuw indien externe themaboom gearchiveerde, uitdovende of uitgedoofde onderwerpen bevat
    if 'extern' in html:
        if '<span class="warn">ARCHIEF' in html['extern']:
            print('*** Let op: de externe themaboom bevat één of meer gearchiveerde onderwerpen! ***')
        if '<span class="warn">UITDOVEND' in html['extern']:
            print('*** Let op: de externe themaboom bevat één of meer uitdovende onderwerpen! ***')
        if '<span class="warn">UITGEDOOFD' in html['extern']:
            print('*** Let op: de externe themaboom bevat één of meer uitgedoofde onderwerpen! ***')

    if export_dates['cat'] != export_dates['ind']:
        print()
        print(f"*** Let op: de exportdatums van {category_tree} ({export_dates['cat']}) en {indicators} ({export_dates['ind']}) zijn niet gelijk. ***")
        print('    Voor een betrouwbaar resultaat moet je beide bestanden op hetzelfde moment exporteren.')

    return  # einde programma

# hoofdprogramma aanroepen als een functie
if __name__ == '__main__':
    main()
